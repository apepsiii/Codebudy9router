"""
Kiro Web Dashboard - FastAPI Backend
"""
from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
import asyncio
import os
import io
from datetime import datetime

# Import database models
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from database import init_db, get_session, Account, Config, ProcessLog

app = FastAPI(title="Kiro Token Generator", version="1.0.0")

# Initialize database on startup
@app.on_event("startup")
async def startup_event():
    init_db()

# Pydantic models for API
class AccountCreate(BaseModel):
    email: str
    password: str
    account_type: str = "login"  # login or register

class AccountBulkCreate(BaseModel):
    accounts: List[AccountCreate]

class ProcessConfig(BaseModel):
    workers: int = 2
    delay: float = 3.0
    visible: bool = False
    manual_mode: bool = False

class AccountResponse(BaseModel):
    id: int
    email: str
    account_type: str
    status: str
    refresh_token: Optional[str] = None
    error_message: Optional[str] = None
    created_at: datetime
    processed_at: Optional[datetime] = None
    injected_to_9router: bool
    injected_at: Optional[datetime] = None
    router_connection_id: Optional[str] = None

    class Config:
        from_attributes = True


class RouterConfig(BaseModel):
    router_url: str = "http://localhost:20128"
    router_password: Optional[str] = None


# ============================================================================
# API ENDPOINTS
# ============================================================================

@app.get("/")
async def root():
    """Serve main dashboard HTML"""
    html_file = os.path.join(os.path.dirname(__file__), "static", "index.html")
    if os.path.exists(html_file):
        with open(html_file, "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    return {"message": "Kiro Web Dashboard API"}


@app.get("/api/stats")
async def get_stats():
    """Get dashboard statistics"""
    session = get_session()
    try:
        total = session.query(Account).count()
        pending = session.query(Account).filter(Account.status == "pending").count()
        processing = session.query(Account).filter(Account.status == "processing").count()
        success = session.query(Account).filter(Account.status == "success").count()
        failed = session.query(Account).filter(Account.status == "failed").count()
        injected = session.query(Account).filter(Account.injected_to_9router == True).count()
        
        return {
            "total": total,
            "pending": pending,
            "processing": processing,
            "success": success,
            "failed": failed,
            "injected": injected,
            "success_rate": round((success / total * 100) if total > 0 else 0, 1)
        }
    finally:
        session.close()


@app.get("/api/accounts", response_model=List[AccountResponse])
async def get_accounts(
    status: Optional[str] = None,
    account_type: Optional[str] = None,
    limit: int = 100,
    offset: int = 0
):
    """Get accounts with optional filters"""
    session = get_session()
    try:
        query = session.query(Account)
        
        if status:
            query = query.filter(Account.status == status)
        if account_type:
            query = query.filter(Account.account_type == account_type)
        
        accounts = query.order_by(Account.created_at.desc()).offset(offset).limit(limit).all()
        return accounts
    finally:
        session.close()


@app.post("/api/accounts", response_model=AccountResponse)
async def create_account(account: AccountCreate):
    """Create single account"""
    session = get_session()
    try:
        # Check if email already exists
        existing = session.query(Account).filter(Account.email == account.email).first()
        if existing:
            raise HTTPException(status_code=400, detail="Email already exists")
        
        db_account = Account(
            email=account.email,
            password=account.password,
            account_type=account.account_type,
            status="pending"
        )
        session.add(db_account)
        session.commit()
        session.refresh(db_account)
        return db_account
    finally:
        session.close()


@app.post("/api/accounts/bulk")
async def create_accounts_bulk(bulk: AccountBulkCreate):
    """Create multiple accounts"""
    session = get_session()
    try:
        created = []
        skipped = []
        
        for account_data in bulk.accounts:
            # Check if email already exists
            existing = session.query(Account).filter(Account.email == account_data.email).first()
            if existing:
                skipped.append(account_data.email)
                continue
            
            db_account = Account(
                email=account_data.email,
                password=account_data.password,
                account_type=account_data.account_type,
                status="pending"
            )
            session.add(db_account)
            created.append(account_data.email)
        
        session.commit()
        
        return {
            "created": len(created),
            "skipped": len(skipped),
            "created_emails": created,
            "skipped_emails": skipped
        }
    finally:
        session.close()


@app.patch("/api/accounts/{account_id}/mark-injected")
async def mark_injected(account_id: int):
    """Mark account as manually injected to 9router"""
    session = get_session()
    try:
        account = session.query(Account).filter(Account.id == account_id).first()
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        account.injected_to_9router = not account.injected_to_9router
        account.injected_at = datetime.utcnow() if account.injected_to_9router else None
        session.commit()
        return {"injected": account.injected_to_9router}
    finally:
        session.close()


@app.post("/api/accounts/reset")
async def reset_accounts(status: Optional[str] = None):
    """Reset accounts back to pending status"""
    session = get_session()
    try:
        query = session.query(Account)
        if status:
            query = query.filter(Account.status == status)
        else:
            query = query.filter(Account.status.in_(["failed", "success"]))
        accounts = query.all()
        for account in accounts:
            account.status = "pending"
            account.error_message = None
        session.commit()
        return {"reset": len(accounts)}
    finally:
        session.close()


@app.delete("/api/accounts/{account_id}")
async def delete_account(account_id: int):
    """Delete account"""
    session = get_session()
    try:
        account = session.query(Account).filter(Account.id == account_id).first()
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        session.delete(account)
        session.commit()
        return {"message": "Account deleted"}
    finally:
        session.close()


@app.delete("/api/accounts")
async def delete_all_accounts(status: Optional[str] = None):
    """Delete all accounts or by status"""
    session = get_session()
    try:
        query = session.query(Account)
        if status:
            query = query.filter(Account.status == status)
        
        count = query.delete()
        session.commit()
        return {"deleted": count}
    finally:
        session.close()


@app.post("/api/process/start")
async def start_processing(config: ProcessConfig, background_tasks: BackgroundTasks):
    """Start processing accounts"""
    session = get_session()
    try:
        # Get pending accounts
        pending_accounts = session.query(Account).filter(Account.status == "pending").all()
        
        if not pending_accounts:
            raise HTTPException(status_code=400, detail="No pending accounts to process")
        
        # Add to background task
        background_tasks.add_task(process_accounts_background, [acc.id for acc in pending_accounts], config)
        
        return {
            "message": "Processing started",
            "accounts": len(pending_accounts),
            "workers": config.workers
        }
    finally:
        session.close()


@app.post("/api/process/stop")
async def stop_processing():
    """Stop processing (placeholder)"""
    # TODO: Implement stop mechanism
    return {"message": "Stop signal sent"}


@app.get("/api/logs/recent")
async def get_recent_logs(limit: int = 30):
    """Get recent logs for all processing accounts"""
    session = get_session()
    try:
        processing_ids = [a.id for a in session.query(Account).filter(Account.status == "processing").all()]
        if not processing_ids:
            return []
        logs = session.query(ProcessLog).filter(
            ProcessLog.account_id.in_(processing_ids)
        ).order_by(ProcessLog.created_at.desc()).limit(limit).all()
        return [{
            "id": log.id,
            "account_id": log.account_id,
            "log_type": log.log_type,
            "message": log.message,
            "created_at": log.created_at.isoformat()
        } for log in logs]
    finally:
        session.close()


@app.get("/api/logs/{account_id}")
async def get_account_logs(account_id: int, limit: int = 50):
    """Get logs for specific account"""
    session = get_session()
    try:
        logs = session.query(ProcessLog).filter(
            ProcessLog.account_id == account_id
        ).order_by(ProcessLog.created_at.desc()).limit(limit).all()
        
        return [{
            "id": log.id,
            "log_type": log.log_type,
            "message": log.message,
            "created_at": log.created_at.isoformat()
        } for log in logs]
    finally:
        session.close()


@app.post("/api/export/tokens")
async def export_tokens():
    """Export all successful tokens to kiro_tokens.txt"""
    session = get_session()
    try:
        success_accounts = session.query(Account).filter(
            Account.status == "success",
            Account.refresh_token.isnot(None)
        ).all()
        
        if not success_accounts:
            raise HTTPException(status_code=404, detail="No successful accounts to export")
        
        output_file = "kiro_tokens.txt"
        with open(output_file, "w", encoding="utf-8") as f:
            for account in success_accounts:
                f.write(f"{account.email}:{account.refresh_token}\n")
        
        return {
            "message": "Tokens exported",
            "file": output_file,
            "count": len(success_accounts)
        }
    finally:
        session.close()


@app.get("/api/export/excel")
async def export_excel(status: Optional[str] = None):
    """Export accounts to Excel file"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    session = get_session()
    try:
        query = session.query(Account)
        if status:
            query = query.filter(Account.status == status)
        accounts = query.order_by(Account.created_at.desc()).all()

        if not accounts:
            raise HTTPException(status_code=404, detail="No accounts found")

        wb = Workbook()
        ws = wb.active
        ws.title = "Kiro Accounts"

        header_fill = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center = Alignment(horizontal="center", vertical="center")
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        headers = ["#", "Email", "Password", "Type", "Status", "Refresh Token", "Injected to 9Router", "Injected At", "Processed At", "Error Message", "Created At"]
        col_widths = [5, 35, 20, 10, 12, 80, 18, 20, 20, 40, 20]

        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 22

        status_colors = {
            "success": "D1FAE5",
            "failed": "FEE2E2",
            "processing": "FEF9C3",
            "pending": "F3F4F6",
        }

        for row_idx, account in enumerate(accounts, 2):
            row_fill = PatternFill(
                start_color=status_colors.get(account.status, "FFFFFF"),
                end_color=status_colors.get(account.status, "FFFFFF"),
                fill_type="solid"
            )
            values = [
                row_idx - 1,
                account.email,
                account.password,
                account.account_type,
                account.status,
                account.refresh_token or "",
                "Yes" if account.injected_to_9router else "No",
                account.injected_at.strftime("%Y-%m-%d %H:%M:%S") if account.injected_at else "",
                account.processed_at.strftime("%Y-%m-%d %H:%M:%S") if account.processed_at else "",
                account.error_message or "",
                account.created_at.strftime("%Y-%m-%d %H:%M:%S") if account.created_at else "",
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill
                cell.border = thin
                cell.alignment = Alignment(vertical="center", wrap_text=(col == 6))
            ws.row_dimensions[row_idx].height = 18

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        summary_ws = wb.create_sheet("Summary")
        total = len(accounts)
        success_count = sum(1 for a in accounts if a.status == "success")
        failed_count = sum(1 for a in accounts if a.status == "failed")
        pending_count = sum(1 for a in accounts if a.status == "pending")
        injected_count = sum(1 for a in accounts if a.injected_to_9router)

        summary_data = [
            ("Total Accounts", total),
            ("Success", success_count),
            ("Failed", failed_count),
            ("Pending", pending_count),
            ("Injected to 9Router", injected_count),
            ("Success Rate", f"{round(success_count / total * 100, 1) if total > 0 else 0}%"),
            ("Export Date", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
        ]
        summary_ws.column_dimensions["A"].width = 25
        summary_ws.column_dimensions["B"].width = 20
        for r, (label, value) in enumerate(summary_data, 1):
            lc = summary_ws.cell(row=r, column=1, value=label)
            lc.font = Font(bold=True)
            lc.border = thin
            vc = summary_ws.cell(row=r, column=2, value=value)
            vc.border = thin

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"kiro_accounts_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    finally:
        session.close()


@app.post("/api/inject/9router")
async def inject_to_9router(config: RouterConfig, background_tasks: BackgroundTasks):
    """Inject all successful tokens to 9router"""
    session = get_session()
    try:
        # Get accounts with tokens that haven't been injected
        accounts = session.query(Account).filter(
            Account.status == "success",
            Account.refresh_token.isnot(None),
            Account.injected_to_9router == False
        ).all()
        
        if not accounts:
            raise HTTPException(status_code=400, detail="No tokens to inject")
        
        # Add to background task
        background_tasks.add_task(inject_tokens_background, 
                                 [acc.id for acc in accounts], 
                                 config.router_url, 
                                 config.router_password)
        
        return {
            "message": "Injection started",
            "accounts": len(accounts)
        }
    finally:
        session.close()


@app.post("/api/inject/9router/{account_id}")
async def inject_single_to_9router(account_id: int, config: RouterConfig):
    """Inject single account token to 9router"""
    session = get_session()
    try:
        account = session.query(Account).filter(Account.id == account_id).first()
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        if not account.refresh_token:
            raise HTTPException(status_code=400, detail="No token available")
        
        # Inject to 9router
        result = await inject_single_token(
            router_url=config.router_url,
            router_password=config.router_password,
            email=account.email,
            refresh_token=account.refresh_token
        )
        
        if result["success"]:
            account.injected_to_9router = True
            account.injected_at = datetime.utcnow()
            account.router_connection_id = result.get("connection_id")
            session.commit()
            
            return {
                "success": True,
                "message": f"Token injected for {account.email}",
                "connection_id": result.get("connection_id")
            }
        else:
            return {
                "success": False,
                "error": result.get("error")
            }
    finally:
        session.close()


@app.get("/api/9router/config")
async def get_9router_config():
    """Get 9router configuration"""
    session = get_session()
    try:
        config = session.query(Config).filter(Config.key == "9router_url").first()
        router_url = config.value if config else "http://localhost:20128"
        
        return {
            "router_url": router_url
        }
    finally:
        session.close()


@app.post("/api/9router/config")
async def save_9router_config(config: RouterConfig):
    """Save 9router configuration"""
    session = get_session()
    try:
        # Save router URL
        url_config = session.query(Config).filter(Config.key == "9router_url").first()
        if url_config:
            url_config.value = config.router_url
            url_config.updated_at = datetime.utcnow()
        else:
            url_config = Config(key="9router_url", value=config.router_url)
            session.add(url_config)
        
        session.commit()
        return {"message": "Configuration saved"}
    finally:
        session.close()


# ============================================================================
# BACKGROUND TASK - Inject Tokens to 9Router
# ============================================================================

async def inject_single_token(router_url: str, router_password: Optional[str], email: str, refresh_token: str):
    """Inject single token to 9router using requests (handles gzip/Cloudflare)"""
    import requests as req_lib

    base_url = router_url.rstrip("/")
    browser_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        "Origin": base_url,
        "Referer": f"{base_url}/",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
    }

    session = req_lib.Session()
    session.headers.update(browser_headers)

    if router_password:
        try:
            resp = session.post(f"{base_url}/api/auth/login", json={"password": router_password}, timeout=15)
            if resp.status_code != 200:
                return {"success": False, "error": f"Login failed: HTTP {resp.status_code}"}
        except Exception as e:
            return {"success": False, "error": f"Login failed: {e}"}

    try:
        resp = session.post(f"{base_url}/api/oauth/kiro/import", json={"refreshToken": refresh_token}, timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            if data.get("success"):
                return {"success": True, "connection_id": data.get("connection", {}).get("id")}
            return {"success": False, "error": "Import failed"}
        return {"success": False, "error": f"HTTP {resp.status_code}: {resp.text[:100]}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


async def inject_tokens_background(account_ids: list, router_url: str, router_password: Optional[str]):
    """Background task to inject tokens to 9router"""
    session = get_session()
    try:
        for account_id in account_ids:
            account = session.query(Account).filter(Account.id == account_id).first()
            if not account or not account.refresh_token:
                continue
            
            # Log
            log = ProcessLog(
                account_id=account_id,
                log_type="info",
                message=f"Injecting token for {account.email} to 9router..."
            )
            session.add(log)
            session.commit()
            
            # Inject
            result = await inject_single_token(
                router_url=router_url,
                router_password=router_password,
                email=account.email,
                refresh_token=account.refresh_token
            )
            
            if result["success"]:
                account.injected_to_9router = True
                account.injected_at = datetime.utcnow()
                account.router_connection_id = result.get("connection_id")
                
                log = ProcessLog(
                    account_id=account_id,
                    log_type="success",
                    message=f"Successfully injected token for {account.email} to 9router"
                )
            else:
                log = ProcessLog(
                    account_id=account_id,
                    log_type="error",
                    message=f"Failed to inject {account.email}: {result.get('error')}"
                )
            
            session.add(log)
            session.commit()
            
            await asyncio.sleep(0.5)  # Small delay between injections
    finally:
        session.close()


# ============================================================================
# BACKGROUND TASK - Process Accounts
# ============================================================================

async def process_accounts_background(account_ids: List[int], config: ProcessConfig):
    """Background task to process accounts using real Playwright automation from main.py"""
    import sys
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    from main import process_account, BROWSER_ARGS
    from playwright.async_api import async_playwright
    from playwright_stealth import Stealth

    session = get_session()
    try:
        accounts_data = []
        for account_id in account_ids:
            account = session.query(Account).filter(Account.id == account_id).first()
            if not account:
                continue
            account.status = "processing"
            log = ProcessLog(
                account_id=account_id,
                log_type="info",
                message=f"Starting processing for {account.email}..."
            )
            session.add(log)
            accounts_data.append((account_id, account.email, account.password, account.account_type))
        session.commit()
    finally:
        session.close()

    headless = not config.visible
    register_mode = False

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=headless,
            args=BROWSER_ARGS + ["--start-maximized"]
        )
        stealth = Stealth()

        for idx, (account_id, email, password, account_type) in enumerate(accounts_data):
            ctx = await browser.new_context(
                permissions=["clipboard-read", "clipboard-write"],
                viewport={"width": 1920, "height": 1080},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
                java_script_enabled=True,
                locale="en-US",
            )
            await ctx.add_init_script("""
                try {
                    Object.defineProperty(navigator, 'webdriver', {get: () => undefined, configurable: true});
                } catch(e) {}
                if (!window.chrome) { window.chrome = { runtime: {} }; }
            """)

            result = await process_account(
                ctx, email, password,
                index=idx + 1,
                total=len(accounts_data),
                worker_id=1,
                register_mode=(account_type == "register"),
                manual_mode=config.manual_mode,
            )
            await ctx.close()

            session = get_session()
            try:
                account = session.query(Account).filter(Account.id == account_id).first()
                if not account:
                    continue

                if result.get("success") and result.get("refresh_token"):
                    account.status = "success"
                    account.refresh_token = result["refresh_token"]
                    account.processed_at = datetime.utcnow()

                    output_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "kiro_tokens.txt")
                    with open(output_file, "a", encoding="utf-8") as f:
                        f.write(f"{email}:{result['refresh_token']}\n")

                    log = ProcessLog(
                        account_id=account_id,
                        log_type="success",
                        message=f"Token captured for {email}: ...{result['refresh_token'][-8:]}"
                    )
                else:
                    account.status = "failed"
                    account.error_message = result.get("error", "Unknown error")
                    log = ProcessLog(
                        account_id=account_id,
                        log_type="error",
                        message=f"Failed to process {email}: {result.get('error', 'Unknown error')}"
                    )

                session.add(log)
                session.commit()
            finally:
                session.close()

            await asyncio.sleep(config.delay)

        await browser.close()


# Mount static files
static_dir = os.path.join(os.path.dirname(__file__), "static")
if os.path.exists(static_dir):
    app.mount("/static", StaticFiles(directory=static_dir), name="static")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=True)
